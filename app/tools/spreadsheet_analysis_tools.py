import os
import logging
import io
import json
from typing import Dict, List, Any, Tuple, Optional
import pandas as pd
from app.services.storage_service import storage_service
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from agents import function_tool

logger = logging.getLogger(__name__)

@function_tool
def analyze_spreadsheet(file_handle: str, analysis_config: str) -> str:
    """
    Performs complex analysis operations on a spreadsheet file.
    
    Args:
        file_handle: The handle of the spreadsheet file to analyze.
        analysis_config: JSON string containing analysis configuration. Format:
            {
                "operations": [
                    {
                        "type": "summary_stats",
                        "target": "B2:B10",
                        "sheet": "Sheet1",
                        "metrics": ["mean", "median", "sum", "min", "max", "count"]
                    },
                    {
                        "type": "filter",
                        "target": "A1:D20",
                        "sheet": "Sheet1", 
                        "condition": {
                            "column": "B",
                            "operator": ">",
                            "value": 100
                        }
                    },
                    {
                        "type": "pivot",
                        "source_data": "A1:D20",
                        "sheet": "Sheet1",
                        "rows": ["A"],
                        "columns": ["C"],
                        "values": ["B"],
                        "aggregation": "sum"
                    },
                    {
                        "type": "extract",
                        "target": "A1:D20",
                        "sheet": "Sheet1",
                        "format": "json"
                    },
                    {
                        "type": "formula_result",
                        "formula": "=SUM(B2:B10)/COUNT(B2:B10)",
                        "sheet": "Sheet1"
                    },
                    {
                        "type": "correlation",
                        "columns": ["B", "C"],
                        "sheet": "Sheet1",
                        "range": "A1:D20"
                    },
                    {
                        "type": "trend_analysis",
                        "x_column": "A",
                        "y_column": "B",
                        "sheet": "Sheet1",
                        "range": "A1:B20"
                    }
                ]
            }
    
    Returns:
        JSON string containing the analysis results.
    """
    logger.info(f"Analyzing spreadsheet '{file_handle}' with specified operations.")
    
    try:
        # Download the file
        print(f"Downloading file '{file_handle}' for analysis.")
        file_bytes = storage_service.download_file(file_handle)
        excel_io = io.BytesIO(file_bytes)
        
        # Parse the analysis configuration
        try:
            config = json.loads(analysis_config)
            operations = config.get("operations", [])
            if not operations:
                return json.dumps({"error": "No analysis operations specified"})
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in analysis configuration: {e}")
            return json.dumps({"error": f"Invalid format for analysis configuration try again with a valid json string: {e}"})
        
        # Load the workbook
        try:
            wb = load_workbook(excel_io, data_only=True)  # data_only=True to get values from formulas
        except Exception as e:
            logger.exception(f"Error reading Excel file '{file_handle}' with openpyxl: {e}")
            return json.dumps({"error": f"Could not read the Excel file: {e}"})
        
        # Process each operation
        results = {}
        
        for i, operation in enumerate(operations):
            op_type = operation.get("type")
            sheet_name = operation.get("sheet", wb.sheetnames[0])
            
            if sheet_name not in wb:
                results[f"operation_{i}"] = {"error": f"Sheet '{sheet_name}' not found"}
                continue
                
            ws = wb[sheet_name]
            
            try:
                if op_type == "summary_stats":
                    target_range = operation.get("target")
                    metrics = operation.get("metrics", ["mean", "sum"])
                    
                    # Extract values from the range
                    values = []
                    for row in ws[target_range]:
                        for cell in row:
                            if isinstance(cell.value, (int, float)):
                                values.append(cell.value)
                    
                    # Calculate statistics
                    stats = {}
                    if "mean" in metrics and values:
                        stats["mean"] = sum(values) / len(values)
                    if "median" in metrics and values:
                        sorted_values = sorted(values)
                        mid = len(sorted_values) // 2
                        if len(sorted_values) % 2 == 0:
                            stats["median"] = (sorted_values[mid-1] + sorted_values[mid]) / 2
                        else:
                            stats["median"] = sorted_values[mid]
                    if "sum" in metrics and values:
                        stats["sum"] = sum(values)
                    if "min" in metrics and values:
                        stats["min"] = min(values)
                    if "max" in metrics and values:
                        stats["max"] = max(values)
                    if "count" in metrics:
                        stats["count"] = len(values)
                    
                    results[f"operation_{i}"] = {
                        "type": "summary_stats",
                        "target": target_range,
                        "results": stats
                    }
                
                elif op_type == "filter":
                    target_range = operation.get("target")
                    condition = operation.get("condition", {})
                    
                    if not condition:
                        results[f"operation_{i}"] = {"error": "No filter condition specified"}
                        continue
                    
                    column_letter = condition.get("column")
                    operator = condition.get("operator")
                    value = condition.get("value")
                    
                    # Get column index
                    col_idx = column_index_from_string(column_letter) - 1
                    
                    # Extract data from range
                    data = []
                    headers = []
                    for row_idx, row in enumerate(ws[target_range]):
                        row_data = [cell.value for cell in row]
                        if row_idx == 0:
                            headers = row_data
                            continue
                        
                        # Apply filter
                        include_row = False
                        cell_value = row_data[col_idx]
                        
                        if operator == ">" and isinstance(cell_value, (int, float)):
                            include_row = cell_value > value
                        elif operator == ">=" and isinstance(cell_value, (int, float)):
                            include_row = cell_value >= value
                        elif operator == "<" and isinstance(cell_value, (int, float)):
                            include_row = cell_value < value
                        elif operator == "<=" and isinstance(cell_value, (int, float)):
                            include_row = cell_value <= value
                        elif operator == "==" or operator == "=":
                            include_row = cell_value == value
                        elif operator == "!=" or operator == "<>":
                            include_row = cell_value != value
                        elif operator == "contains" and isinstance(cell_value, str):
                            include_row = value in cell_value
                        
                        if include_row:
                            data.append(row_data)
                    
                    results[f"operation_{i}"] = {
                        "type": "filter",
                        "headers": headers,
                        "filtered_data": data,
                        "count": len(data)
                    }
                
                elif op_type == "extract":
                    target_range = operation.get("target")
                    format_type = operation.get("format", "json")
                    
                    # Extract data from range
                    data = []
                    headers = []
                    for row_idx, row in enumerate(ws[target_range]):
                        row_data = [cell.value for cell in row]
                        if row_idx == 0:
                            headers = row_data
                            continue
                        
                        data.append(row_data)
                    
                    if format_type == "json":
                        # Convert to list of dictionaries
                        json_data = []
                        for row in data:
                            row_dict = {}
                            for i, header in enumerate(headers):
                                if i < len(row):
                                    row_dict[header] = row[i]
                            json_data.append(row_dict)
                        
                        results[f"operation_{i}"] = {
                            "type": "extract",
                            "data": json_data
                        }
                    else:
                        results[f"operation_{i}"] = {
                            "type": "extract",
                            "headers": headers,
                            "data": data
                        }
                
                elif op_type == "formula_result":
                    formula = operation.get("formula")
                    
                    # Create a temporary cell with the formula
                    temp_cell = ws.cell(row=1, column=1)
                    original_value = temp_cell.value
                    temp_cell.value = formula
                    
                    # Save and reload to calculate the formula
                    temp_io = io.BytesIO()
                    wb.save(temp_io)
                    temp_io.seek(0)
                    
                    # Load with data_only=True to get the calculated value
                    temp_wb = load_workbook(temp_io, data_only=True)
                    temp_ws = temp_wb[sheet_name]
                    result_value = temp_ws.cell(row=1, column=1).value
                    
                    # Restore original value
                    temp_cell.value = original_value
                    
                    results[f"operation_{i}"] = {
                        "type": "formula_result",
                        "formula": formula,
                        "result": result_value
                    }
                
                elif op_type == "correlation":
                    columns = operation.get("columns", [])
                    range_str = operation.get("range")
                    
                    if len(columns) != 2:
                        results[f"operation_{i}"] = {"error": "Correlation requires exactly 2 columns"}
                        continue
                    
                    col1_letter, col2_letter = columns
                    col1_idx = column_index_from_string(col1_letter) - 1
                    col2_idx = column_index_from_string(col2_letter) - 1
                    
                    # Extract data from range
                    col1_data = []
                    col2_data = []
                    
                    for row_idx, row in enumerate(ws[range_str]):
                        if row_idx == 0:  # Skip header
                            continue
                            
                        row_data = [cell.value for cell in row]
                        
                        if col1_idx < len(row_data) and col2_idx < len(row_data):
                            val1 = row_data[col1_idx]
                            val2 = row_data[col2_idx]
                            
                            if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
                                col1_data.append(val1)
                                col2_data.append(val2)
                    
                    # Calculate correlation if we have data
                    if col1_data and col2_data and len(col1_data) == len(col2_data):
                        n = len(col1_data)
                        sum_x = sum(col1_data)
                        sum_y = sum(col2_data)
                        sum_xy = sum(x * y for x, y in zip(col1_data, col2_data))
                        sum_x2 = sum(x * x for x in col1_data)
                        sum_y2 = sum(y * y for y in col2_data)
                        
                        # Pearson correlation coefficient
                        numerator = n * sum_xy - sum_x * sum_y
                        denominator = ((n * sum_x2 - sum_x ** 2) * (n * sum_y2 - sum_y ** 2)) ** 0.5
                        
                        if denominator != 0:
                            correlation = numerator / denominator
                        else:
                            correlation = None
                            
                        results[f"operation_{i}"] = {
                            "type": "correlation",
                            "columns": columns,
                            "correlation": correlation,
                            "sample_size": n
                        }
                    else:
                        results[f"operation_{i}"] = {
                            "type": "correlation",
                            "error": "Insufficient numeric data for correlation"
                        }
                
                elif op_type == "trend_analysis":
                    x_column = operation.get("x_column")
                    y_column = operation.get("y_column")
                    range_str = operation.get("range")
                    
                    x_col_idx = column_index_from_string(x_column) - 1
                    y_col_idx = column_index_from_string(y_column) - 1
                    
                    # Extract data from range
                    x_data = []
                    y_data = []
                    
                    for row_idx, row in enumerate(ws[range_str]):
                        if row_idx == 0:  # Skip header
                            continue
                            
                        row_data = [cell.value for cell in row]
                        
                        if x_col_idx < len(row_data) and y_col_idx < len(row_data):
                            x_val = row_data[x_col_idx]
                            y_val = row_data[y_col_idx]
                            
                            # For trend analysis, we need numeric y values
                            if isinstance(y_val, (int, float)):
                                # x can be date or numeric
                                x_data.append(x_val)
                                y_data.append(y_val)
                    
                    # Calculate linear regression if we have data
                    if x_data and y_data and len(x_data) == len(y_data):
                        # Convert dates to numeric if needed
                        numeric_x_data = []
                        for x in x_data:
                            if isinstance(x, (int, float)):
                                numeric_x_data.append(x)
                            else:
                                # Try to treat as ordinal (position in sequence)
                                numeric_x_data.append(len(numeric_x_data) + 1)
                        
                        n = len(numeric_x_data)
                        sum_x = sum(numeric_x_data)
                        sum_y = sum(y_data)
                        sum_xy = sum(x * y for x, y in zip(numeric_x_data, y_data))
                        sum_x2 = sum(x * x for x in numeric_x_data)
                        
                        # Linear regression: y = mx + b
                        try:
                            m = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x ** 2)
                            b = (sum_y - m * sum_x) / n
                            
                            # Calculate R-squared
                            y_mean = sum_y / n
                            ss_total = sum((y - y_mean) ** 2 for y in y_data)
                            ss_residual = sum((y - (m * x + b)) ** 2 for x, y in zip(numeric_x_data, y_data))
                            
                            if ss_total != 0:
                                r_squared = 1 - (ss_residual / ss_total)
                            else:
                                r_squared = None
                                
                            # Predict next value
                            next_x = max(numeric_x_data) + 1
                            next_y_prediction = m * next_x + b
                            
                            results[f"operation_{i}"] = {
                                "type": "trend_analysis",
                                "slope": m,
                                "intercept": b,
                                "r_squared": r_squared,
                                "sample_size": n,
                                "next_value_prediction": next_y_prediction
                            }
                        except Exception as e:
                            results[f"operation_{i}"] = {
                                "type": "trend_analysis",
                                "error": f"Error calculating trend: {str(e)}"
                            }
                    else:
                        results[f"operation_{i}"] = {
                            "type": "trend_analysis",
                            "error": "Insufficient data for trend analysis"
                        }
                
                elif op_type == "pivot":
                    source_data = operation.get("source_data")
                    row_cols = operation.get("rows", [])
                    col_cols = operation.get("columns", [])
                    val_cols = operation.get("values", [])
                    aggregation = operation.get("aggregation", "sum")
                    
                    if not row_cols or not val_cols:
                        results[f"operation_{i}"] = {"error": "Pivot requires at least rows and values"}
                        continue
                    
                    # Extract data and headers from range
                    data = []
                    headers = []
                    
                    for row_idx, row in enumerate(ws[source_data]):
                        row_data = [cell.value for cell in row]
                        if row_idx == 0:
                            headers = row_data
                            continue
                        
                        data.append(row_data)
                    
                    # Convert column letters to indices
                    row_indices = [column_index_from_string(col) - 1 for col in row_cols]
                    col_indices = [column_index_from_string(col) - 1 for col in col_cols]
                    val_indices = [column_index_from_string(col) - 1 for col in val_cols]
                    
                    # Build pivot table
                    pivot_data = {}
                    for row in data:
                        # Create row key (can be multiple columns)
                        row_key = tuple(row[idx] if idx < len(row) else None for idx in row_indices)
                        
                        # Create column key (can be multiple columns)
                        col_key = tuple(row[idx] if idx < len(row) else None for idx in col_indices) if col_indices else ("Value",)
                        
                        # Get values
                        for val_idx in val_indices:
                            if val_idx < len(row) and isinstance(row[val_idx], (int, float)):
                                value = row[val_idx]
                                
                                if row_key not in pivot_data:
                                    pivot_data[row_key] = {}
                                
                                if col_key not in pivot_data[row_key]:
                                    pivot_data[row_key][col_key] = []
                                
                                pivot_data[row_key][col_key].append(value)
                    
                    # Apply aggregation
                    for row_key in pivot_data:
                        for col_key in pivot_data[row_key]:
                            values = pivot_data[row_key][col_key]
                            
                            if aggregation == "sum":
                                pivot_data[row_key][col_key] = sum(values)
                            elif aggregation == "avg" or aggregation == "mean":
                                pivot_data[row_key][col_key] = sum(values) / len(values)
                            elif aggregation == "min":
                                pivot_data[row_key][col_key] = min(values)
                            elif aggregation == "max":
                                pivot_data[row_key][col_key] = max(values)
                            elif aggregation == "count":
                                pivot_data[row_key][col_key] = len(values)
                    
                    # Convert to a more JSON-friendly format
                    pivot_result = []
                    unique_col_keys = set()
                    
                    for row_key in pivot_data:
                        unique_col_keys.update(pivot_data[row_key].keys())
                    
                    unique_col_keys = sorted(list(unique_col_keys))
                    
                    for row_key in pivot_data:
                        row_dict = {}
                        
                        # Add row identifiers
                        for i, idx in enumerate(row_indices):
                            if idx < len(headers):
                                row_dict[headers[idx]] = row_key[i]
                        
                        # Add values for each column
                        for col_key in unique_col_keys:
                            col_name = "_".join(str(x) for x in col_key if x is not None)
                            row_dict[col_name] = pivot_data[row_key].get(col_key, None)
                        
                        pivot_result.append(row_dict)
                    
                    results[f"operation_{i}"] = {
                        "type": "pivot",
                        "pivot_data": pivot_result
                    }
                
                else:
                    results[f"operation_{i}"] = {"error": f"Unsupported operation type: {op_type}"}
            
            except Exception as e:
                logger.exception(f"Error processing operation '{op_type}': {e}")
                results[f"operation_{i}"] = {"error": f"Error processing operation '{op_type}': {str(e)}"}
        
        return json.dumps(results)
    
    except FileNotFoundError:
        logger.error(f"Could not find file with handle '{file_handle}' for analysis.")
        return json.dumps({"error": f"File not found: {file_handle}"})
    except Exception as e:
        logger.exception(f"Error analyzing spreadsheet '{file_handle}': {e}")
        return json.dumps({"error": f"Error analyzing spreadsheet: {str(e)}"})

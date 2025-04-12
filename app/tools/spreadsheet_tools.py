import os
import logging
import io
import json
from typing import Dict, List, Any, Tuple, Optional
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment
from app.services.storage_service import storage_service
from agents import function_tool
from app.services.gemini_service import GeminiService
from app.services.script_execution_service import ScriptExecutionService
from app.tools.spreadsheet_utils import read_excel_all, get_excel_bytes

logger = logging.getLogger(__name__)

@function_tool
def ponder_spreadsheet_request(request_description: str, points_to_consider: str) -> str:
    """
    A tool for the spreadsheet agent to think through a request before taking action.
    This allows for more careful planning and consideration of the appropriate approach.
    
    Args:
        request_description: A brief description of the client's request
        points_to_consider: Key points to consider when handling this request
        
    Returns:
        Guidance on how to proceed with the request
    """
    print(f"Spreadsheet agent pondering request: {request_description}")
    print(f"Points to consider: {points_to_consider}")
    
    # For now, this just echoes back the points with some generic guidance
    response = {
        "analysis": points_to_consider,
        "recommendation": "Based on your analysis, proceed with the appropriate spreadsheet tool calls to fulfill the request. Remember to use read_file_content first if working with an existing spreadsheet, create_spreadsheet for new spreadsheets, or modify_spreadsheet to make changes to existing files."
    }
    
    return json.dumps(response)

def apply_cell_format(cell, format_spec):
    """Apply formatting to a cell based on the format specification."""
    # Font formatting
    if format_spec.get("bold"):
        if cell.font:
            cell.font = Font(bold=True, name=cell.font.name, size=cell.font.size)
        else:
            cell.font = Font(bold=True)
    
    # Background color
    if format_spec.get("bg_color"):
        # Ensure color is in aRGB format (8 characters with FF alpha channel)
        color = format_spec["bg_color"]
        # If it's a 6-character hex, add FF alpha channel prefix
        if len(color) == 6 and all(c in "0123456789ABCDEFabcdef" for c in color):
            color = "FF" + color
        # If it doesn't start with FF and is 6 chars, add the alpha
        elif len(color) == 6:
            color = "FF" + color
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # Number format
    if format_spec.get("number_format"):
        cell.number_format = format_spec["number_format"]
    
    # Alignment
    if format_spec.get("align"):
        align_type = format_spec["align"]
        if align_type == "center":
            cell.alignment = Alignment(horizontal="center")
        elif align_type == "right":
            cell.alignment = Alignment(horizontal="right")
        elif align_type == "left":
            cell.alignment = Alignment(horizontal="left")

@function_tool
def create_spreadsheet(filename: str, spreadsheet_data: str) -> str:
    """
    Creates a new spreadsheet file (.xlsx) with the given data, formulas, and formatting.

    Args:
        filename: The desired name for the spreadsheet file (including extension).
        spreadsheet_data: JSON string containing sheet data, formulas, and formatting. Format:
            {
                "sheets": [
                    {
                        "name": "Sheet1",
                        "data": [
                            ["Header1", "Header2", "Header3"],
                            ["Value1", 10, "=B2*2"],
                            ["Value2", 20, "=B3*2"]
                        ],
                        "column_widths": {"A": 15, "B": 10},
                        "formats": [
                            {"range": "A1:C1", "bold": true, "bg_color": "FFCCCCCC"},
                            {"range": "B2:B3", "number_format": "#,##0.00"}
                        ]
                    }
                ]
            }

    Returns:
        The file handle of the created file.
    """
    logger.info(f"Attempting to create spreadsheet '{filename}'.")
    # Ensure filename has a valid extension, default to .xlsx if not
    if not filename.lower().endswith('.xlsx'):
        logger.warning(f"Filename '{filename}' lacks .xlsx extension. Appending '.xlsx'.")
        filename += '.xlsx'

    try:
        print(f"Creating spreadsheet '{filename}' with advanced formatting")
        
        # Parse the JSON input
        try:
            spec = json.loads(spreadsheet_data)
        except json.JSONDecodeError:
            # Fallback to CSV format if JSON parsing fails
            print("JSON parsing failed, falling back to CSV format")
            csv_data_io = io.StringIO(spreadsheet_data)
            df = pd.read_csv(csv_data_io)
            
            # Create a workbook with the DataFrame
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name).font = Font(bold=True)
                
            # Write data
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
        else:
            # Process the JSON specification
            wb = openpyxl.Workbook()
            
            # Remove the default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Process each sheet in the specification
            for sheet_idx, sheet_spec in enumerate(spec.get("sheets", [])):
                sheet_name = sheet_spec.get("name", f"Sheet{sheet_idx+1}")
                ws = wb.create_sheet(title=sheet_name)
                
                # Add data and formulas
                data = sheet_spec.get("data", [])
                for row_idx, row_data in enumerate(data, 1):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        
                        # Handle formulas (starting with =)
                        if isinstance(cell_value, str) and cell_value.startswith('='):
                            cell.value = cell_value
                        else:
                            cell.value = cell_value
                
                # Set column widths
                for col_letter, width in sheet_spec.get("column_widths", {}).items():
                    ws.column_dimensions[col_letter].width = width
                
                # Apply formats
                for format_spec in sheet_spec.get("formats", []):
                    cell_range = format_spec.get("range")
                    if cell_range:
                        for cell in ws[cell_range]:
                            # Apply cell formatting
                            if isinstance(cell, tuple):
                                for c in cell:
                                    apply_cell_format(c, format_spec)
                            else:
                                apply_cell_format(cell, format_spec)
        
        # Save to bytes
        output_excel_io = io.BytesIO()
        wb.save(output_excel_io)
        excel_bytes = output_excel_io.getvalue()

        # Upload the file bytes using the storage service
        file_handle = storage_service.upload_file(filename, excel_bytes)
        logger.info(f"Successfully created spreadsheet '{filename}' with handle '{file_handle}'.")
        return file_handle

    except Exception as e:
        print(f"Error creating spreadsheet '{filename}': {e}")
        logger.exception(f"Error creating spreadsheet '{filename}': {e}")
        return f"Error creating spreadsheet '{filename}': {e}"

@function_tool
def modify_spreadsheet(file_handle: str, modifications: str) -> str:
    """
    Modifies an existing spreadsheet file identified by its handle.

    Args:
        file_handle: The handle of the spreadsheet file to modify.
        modifications: A JSON string containing a list of operations to perform on the spreadsheet.
          Example: [{"operation": "update_cell", "cell": "B5", "value": "42"}]
          
          Supported operations:
          - update_cell: Updates a single cell value
          - add_row: Adds a new row with the provided data
          - delete_row: Deletes a row at the specified index
          - clear_range: Clears all values in the specified range
          - set_formula: Sets a formula in the specified cell
          - apply_basic_style: Applies basic styling to a range of cells

    Returns:
        The file handle of the modified spreadsheet.
    """
    logger.info(f"Attempting to modify spreadsheet '{file_handle}'. Modifications: {modifications}")

    try:
        # 1. Download the existing file
        print(f"Downloading file '{file_handle}' for modification.")
        file_bytes = storage_service.download_file(file_handle)
        excel_io = io.BytesIO(file_bytes)

        # 2. Parse the JSON modifications string
        try:
            ops = json.loads(modifications)
            if not isinstance(ops, list):
                raise ValueError("Modifications JSON must be a list of operations.")
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in modifications for '{file_handle}': {e}")
            return f"Error: Invalid format for modifications - must be a valid JSON list. {e}"
        except ValueError as e:
             logger.error(f"Invalid JSON structure for '{file_handle}': {e}")
             return f"Error: {e}"

        # 3. Load the workbook using openpyxl
        try:
            wb = load_workbook(excel_io)
        except Exception as e:
             logger.exception(f"Error reading Excel file '{file_handle}' with openpyxl: {e}")
             return f"Error: Could not read the Excel file '{file_handle}'. Is it a valid .xlsx file? {e}"

        # 4. Apply operations using openpyxl
        sheet_names = wb.sheetnames
        if not sheet_names:
             return f"Error: Excel file '{file_handle}' appears to be empty or has no sheets."

        for op in ops:
            if not isinstance(op, dict) or 'operation' not in op:
                logger.warning(f"Skipping invalid operation item: {op}")
                continue

            op_type = op['operation']
            # Default to first sheet if not specified
            sheet_name = op.get('sheet', sheet_names[0]) 

            if sheet_name not in wb:
                logger.error(f"Sheet '{sheet_name}' not found in '{file_handle}'. Available: {sheet_names}")
                return f"Error: Sheet '{sheet_name}' not found in the spreadsheet."
            
            ws = wb[sheet_name] # Get the openpyxl worksheet

            try:
                # --- Reimplementing operations using openpyxl --- 
                if op_type == 'update_cell':
                    cell = op['cell']
                    value = op['value']
                    ws[cell] = value # Direct assignment
                    logger.info(f"Updated cell {cell} in sheet '{sheet_name}' to '{value}'.")

                elif op_type == 'add_row':
                    data = op['data']
                    ws.append(data) # Append row
                    logger.info(f"Added row {data} to sheet '{sheet_name}'.")

                elif op_type == 'delete_row':
                    # openpyxl row indices are 1-based
                    row_index_0based = op['row_index'] 
                    if not isinstance(row_index_0based, int) or row_index_0based < 0:
                        raise ValueError("row_index must be a non-negative integer.")
                    ws.delete_rows(row_index_0based + 1, 1)
                    logger.info(f"Deleted row at 0-based index {row_index_0based} from sheet '{sheet_name}'.")

                elif op_type == 'clear_range':
                    range_str = op['range']
                    for row in ws[range_str]:
                        for cell in row:
                            cell.value = None # Clear cell value
                    logger.info(f"Cleared range {range_str} in sheet '{sheet_name}'.")

                # --- New Operations ---
                elif op_type == 'set_formula':
                    cell = op['cell']
                    formula = op['formula']
                    if not formula.startswith('='):
                        formula = '=' + formula # Ensure it's treated as a formula
                    ws[cell] = formula
                    logger.info(f"Set formula in cell {cell} of sheet '{sheet_name}' to '{formula}'.")

                elif op_type == 'apply_basic_style':
                    range_str = op['range']
                    style_dict = op.get('style', {})
                    is_bold = style_dict.get('bold', False)
                    bg_color = style_dict.get('bg_color')
                    number_format = style_dict.get('number_format')
                    
                    for row in ws[range_str]:
                        for cell in row:
                            if is_bold:
                                cell.font = Font(bold=True)
                            if bg_color:
                                # Ensure color is in aRGB format (8 characters with FF alpha channel)
                                color = bg_color
                                if len(color) == 6 and all(c in "0123456789ABCDEFabcdef" for c in color):
                                    color = "FF" + color
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                            if number_format:
                                cell.number_format = number_format
                                
                    logger.info(f"Applied style to range {range_str} in sheet '{sheet_name}'.")
                
                else:
                    logger.warning(f"Unsupported operation type '{op_type}' specified.")
                    return f"Error: Unsupported operation type '{op_type}'."

            except KeyError as e:
                logger.error(f"Missing required key for operation '{op_type}': {e}")
                return f"Error: Missing data for operation '{op_type}': {e}."
            except Exception as e:
                logger.exception(f"Error processing operation '{op_type}' on sheet '{sheet_name}': {e}")
                return f"Unexpected error during operation '{op_type}' on sheet '{sheet_name}': {e}"

        # 5. Save the modified workbook
        output_excel_io = io.BytesIO()
        wb.save(output_excel_io)
        excel_bytes = output_excel_io.getvalue()

        # Upload the modified file
        # Use the same file handle to replace the original file
        storage_service.upload_file(file_handle, excel_bytes)
        logger.info(f"Successfully modified and saved spreadsheet '{file_handle}'.")
        return file_handle
    
    except FileNotFoundError:
        logger.error(f"Could not find file with handle '{file_handle}' for modification.")
        return f"Error: File not found: {file_handle}"
    except Exception as e:
        logger.exception(f"Error modifying spreadsheet '{file_handle}': {e}")
        return f"Error modifying spreadsheet: {e}"

@function_tool
def magic_enhance_excel(file_handle: str, instructions: str = "") -> Dict[str, Any]:
    """Uses AI to analyze and enhance Excel file structure, formatting, and data quality"""
    try:
        # Setup base script template without extra indentation
        script_template = """try:
    # Read the input Excel file
    sheets_data = read_excel_all(input_data['input_handle'])
    enhanced_sheets = {{}}
    changes_made = []

    # Process each sheet
    for sheet_name, df in sheets_data.items():
        # Make a copy of the dataframe for enhancement
        enhanced_df = df.copy()
        
        # Apply enhancements based on analysis
{enhancement_code}
        
        # Store enhanced sheet
        enhanced_sheets[sheet_name] = enhanced_df
        
    # Save enhanced workbook
    output_handle = write_excel_all(enhanced_sheets, f"enhanced_{{input_data['input_handle']}}")
    
    # Return results
    output = {{
        'output_handle': output_handle,
        'changes_made': changes_made,
        'sheets_processed': list(sheets_data.keys())
    }}

except Exception as e:
    output = {{'error': str(e)}}"""

        print("Enhancing Excel file...")


        # Get analysis results first
        file_bytes = get_excel_bytes(file_handle)
        sheets_data = read_excel_all(file_handle)
        wb = load_workbook(io.BytesIO(file_bytes))
        
        # Build analysis prompt with actual data context
        analysis_prompt = f"""
        Generate Python code for Excel enhancement operations.
        The code will be inserted into a template that already handles file loading and saving.
        Focus on generating the enhancement operations only.

        Instructions from user: {instructions}

        Available data operations:
        - DataFrame operations (using pandas)
        - Numeric operations (using numpy)
        - String operations
        - Basic statistical operations
        - Scikit-learn operations 

        Sheet names available: {list(sheets_data.keys())}
        Sample data structure: {json.dumps({name: df.head(3).to_dict() for name, df in sheets_data.items()})}

        Generate enhancement code that:
        1. Uses only pandas DataFrame operations
        2. Tracks changes in the changes_made list
        3. Handles errors gracefully
        4. Returns enhanced DataFrames
        5. Uses the provided script template for context
        6. Manipulate sheet reformat texts, update columns.
        7. Add new columns based on analysis
        8. Add visualization charts
        9. Add colors to text, aand formulas
        10.UUs pyhon to perrom any traansformation a usr can do o an excel sheet
        
        Return ONLY the enhancement code, no imports or file operations.
        """
        
        gemini = GeminiService()
        enhancement_result = gemini.analyze_text(analysis_prompt, "Generate Enhancement Code")
        if not enhancement_result.get("success"):
            raise Exception(f"Failed to generate enhancement code: {enhancement_result.get('error')}")

        # Extract and clean the code
        enhancement_code = enhancement_result["text"]
        if "```python" in enhancement_code:
            enhancement_code = enhancement_code.split("```python")[1].split("```")[0]
        enhancement_code = enhancement_code.strip()
        
        # Indent the enhancement code
        enhancement_code = "\n".join("        " + line for line in enhancement_code.splitlines())
        
        # Create final script
        final_script = script_template.format(enhancement_code=enhancement_code)
        
        # Execute the enhancement script
        script_executor = ScriptExecutionService()
        execution_result = script_executor.execute_script(
            final_script,
            {"input_handle": file_handle}
        )
        
        if not execution_result.get("success"):
            raise Exception(f"Enhancement execution failed: {execution_result.get('error')}")
            
        return {
            "success": True,
            "enhanced_handle": execution_result["output"].get("output_handle"),
            "changes_made": execution_result["output"].get("changes_made", [])
        }

    except Exception as e:
        print(f"Excel enhancement failed: {e}")
        logger.exception(f"Excel enhancement failed: {e}")
        return {
            "success": False,
            "error": str(e)
        }

def _extract_json_from_response(response_text: str) -> dict:
    """Helper function to extract and parse JSON from Gemini response"""
    try:
        # Clean up the response text
        text = response_text.strip()
        
        # Remove markdown headers
        text = '\n'.join(line for line in text.split('\n') 
                        if not line.strip().startswith('#'))
        
        # Case 1: Response is already valid JSON
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            pass
            
        # Case 2: Find JSON content between code blocks
        if "```" in text:
            # Split by code blocks and try each block
            blocks = text.split("```")
            for block in blocks:
                # Remove language identifier if present
                if "\n" in block:
                    block = block.split("\n", 1)[1]
                block = block.strip()
                if not block:
                    continue
                try:
                    return json.loads(block)
                except json.JSONDecodeError:
                    continue
        
        # Case 3: Try to find JSON-like content
        import re
        json_pattern = r'\{[\s\S]*\}'
        matches = re.findall(json_pattern, text)
        for match in matches:
            try:
                return json.loads(match)
            except json.JSONDecodeError:
                continue
            
        raise ValueError("No valid JSON content found in response")
            
    except Exception as e:
        logger.error(f"Failed to parse JSON response: {text}")
        logger.error(f"Parse error: {str(e)}")
        # Return a basic valid JSON structure
        return {
            "findings": ["Error parsing analysis results"],
            "data_quality_issues": [],
            "potential_insights": [],
            "recommended_deep_dives": [],
            "suggested_visualizations": [],
            "next_analysis_steps": []
        }

@function_tool
def smart_spreadsheet_analysis(file_handle: str, depth: str = "standard") -> Dict[str, Any]:
    """
    Performs intelligent spreadsheet analysis with feedback loop and report generation.
    
    Args:
        file_handle: The handle of the Excel file to analyze
        depth: Analysis depth ("quick", "standard", "deep")
        
    Returns:
        Dictionary containing analysis results, report handle, and any additional outputs
    """
    try:
        results = {"analysis_rounds": []}
        
        # Initial data load and preparation
        file_bytes = storage_service.download_file(file_handle)
        sheets_data = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
        wb = load_workbook(io.BytesIO(file_bytes))
        
        # Collect metadata for analysis
        sheets_meta = {}
        for sheet_name, df in sheets_data.items():
            ws = wb[sheet_name]
            meta = {
                "headers": list(df.columns),
                "sample_data": df.head(5).to_dict('records'),
                "statistics": {
                    "rows": len(df),
                    "columns": len(df.columns),
                    "missing_values": df.isnull().sum().to_dict(),
                    "numeric_columns": df.select_dtypes(include=['number']).columns.tolist(),
                    "text_columns": df.select_dtypes(include=['object']).columns.tolist()
                },
                "formulas": [
                    {"cell": cell.coordinate, "formula": cell.value}
                    for row in ws.iter_rows()
                    for cell in row
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('=')
                ]
            }
            sheets_meta[sheet_name] = meta

        gemini = GeminiService()
        
        # Initial Analysis Round
        initial_prompt = f"""
        Perform initial spreadsheet analysis:
        
        Data Structure:
        {json.dumps(sheets_meta, indent=2)}
        
        Analyze:
        1. Data quality and completeness
        2. Structural patterns and relationships
        3. Potential insights and anomalies
        4. Areas needing deeper investigation
        
        Return in JSON:
        {{
            "findings": [],
            "data_quality_issues": [],
            "potential_insights": [],
            "recommended_deep_dives": [],
            "suggested_visualizations": [],
            "next_analysis_steps": []
        }}
        """
        
        initial_result = gemini.analyze_text(initial_prompt, "Initial Analysis")
        if not initial_result.get("success"):
            raise Exception("Initial analysis failed")

        print(f"Initial analysis result: {initial_result['text']}")
            
        try:
            initial_analysis = _extract_json_from_response(initial_result["text"])
        except ValueError as e:
            logger.error(f"Failed to parse initial analysis JSON: {str(e)}")
            raise Exception(f"Failed to parse initial analysis JSON: {str(e)}")
        results["analysis_rounds"].append({"type": "initial", "findings": initial_analysis})
        
        # Feedback Loop - Deep Dive Analysis
        for deep_dive in initial_analysis.get("recommended_deep_dives", [])[:3]:
            dive_prompt = f"""
            Perform deep dive analysis on: {deep_dive}
            
            Previous Findings:
            {json.dumps(initial_analysis, indent=2)}
            
            Available Data:
            {json.dumps(sheets_meta, indent=2)}
            
            Focus on:
            1. Detailed statistical analysis
            2. Pattern identification
            3. Actionable insights
            4. Specific recommendations
            
            Return analysis in this JSON format:
            {{
                "deep_dive_focus": "{deep_dive}",
                "findings": [],
                "statistical_analysis": {{
                    "key_metrics": [],
                    "patterns": [],
                    "outliers": []
                }},
                "recommendations": [],
                "next_steps": []
            }}
            """
            
            dive_result = gemini.analyze_text(dive_prompt, f"Deep Dive: {deep_dive}")
            if dive_result.get("success"):
                try:
                    raw_text = dive_result["text"].strip()
                    # Remove any markdown headers
                    if raw_text.startswith('#'):
                        raw_text = '\n'.join(line for line in raw_text.split('\n') 
                                           if not line.strip().startswith('#'))
                    
                    findings = _extract_json_from_response(raw_text)
                    if not isinstance(findings, dict):
                        findings = {
                            "deep_dive_focus": deep_dive,
                            "findings": ["Error: Analysis returned invalid format"],
                            "statistical_analysis": {
                                "key_metrics": [],
                                "patterns": [],
                                "outliers": []
                            },
                            "recommendations": [],
                            "next_steps": []
                        }
                    
                    results["analysis_rounds"].append({
                        "type": "deep_dive",
                        "focus": deep_dive,
                        "findings": findings
                    })
                except Exception as e:
                    logger.warning(f"Failed to parse deep dive analysis for {deep_dive}: {str(e)}")
                    continue
        
        # Generate Comprehensive Report
        report_prompt = f"""
        Create detailed markdown report from multiple analysis rounds:
        {json.dumps(results["analysis_rounds"], indent=2)}
        
        Include:
        1. Executive Summary
        2. Key Findings and Insights
        3. Data Quality Assessment
        4. Detailed Analysis by Area
        5. Visualizations Recommendations
        6. Action Items and Recommendations
        7. Technical Appendix
        
        Use professional markdown formatting with:
        - Clear section hierarchy
        - Linked table of content
        - Focus on insights and not banal summaries.
        - Visualizations (charts, graphs) where applicable
        - Use the full markdown syntax for tables, lists, and code blocks
        - Use headings and subheadings for organization
        - Tables for structured data
        - Code blocks for technical details
        """
        
        report_result = gemini.analyze_text(report_prompt, "Generate Final Report")
        if not report_result.get("success"):
            raise Exception("Report generation failed")
            
        # Save report
        report_handle = storage_service.upload_file(
            f"smart_analysis_{file_handle.split('/')[-1]}.md",
            report_result.get("text", "").encode('utf-8')
        )
        
        results.update({
            "success": True,
            "report_handle": report_handle,
            "summary": initial_analysis.get("findings", []),
            "recommendations": initial_analysis.get("next_analysis_steps", [])
        })
        
        return results

    except Exception as e:
        logger.exception(f"Smart spreadsheet analysis failed: {e}")
        return {
            "success": False,
            "error": str(e)
        }
